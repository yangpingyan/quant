#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Created on 2016-8-25

@author: YPY
"""


import json
import argparse
import time
import datetime
import smtplib
from email.mime.text import MIMEText  
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from email import encoders
import os
import mimetypes

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

import tushare as ts

def file2dict(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)

def getprice_hs300():    
    df = ts.get_realtime_quotes('hs300')
    price = df['price'][0]
    return price


def modify1002excel(filename = 'c:/1.xlsx', assets=None, hs300price=None, fundunit=None):    
    wb = load_workbook(filename = filename)
    ws = wb['基金资产明细']
    nrows = len(list(ws.rows))
    ncolumns = len(list(ws.columns))
    now = datetime.datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    newrow = nrows+1    
    
    ws.cell(row = newrow, column = 9).value = hs300price
    ws.cell(row = newrow, column = 8).value = assets
    ws.cell(row = newrow, column = 7).value = 0
    ws.cell(row = newrow, column = 6).value = 0
    ws.cell(row = newrow, column = 5).value = assets
    ws.cell(row = newrow, column = 4).value = fundunit
    ws.cell(row = newrow, column = 3).value = hs300price/ws.cell(row = 2, column = 9).value
    ws.cell(row = newrow, column = 2).value = ws.cell(row = newrow, column = 5).value/ws.cell(row = newrow, column = 4).value
    ws.cell(row = newrow, column = 1).value = now
    
   
    for cx in range(1, ncolumns+1) :
        ws.cell(row = newrow, column = cx).number_format = ws.cell(row = nrows, column = cx).number_format        

    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=365)
    
    # Chart with date axis
    c2 = LineChart()
    c2.title = "擎天柱基金净值走势图"
    c2.style = 2
    # c2.y_axis.title = "Size"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'm-d'
    c2.x_axis.majorTimeUnit = "months"
  
    # c2.x_axis.title = "Date"
    
    c2.add_data(data, titles_from_data=True)
    dates = Reference(ws, min_col=1, min_row=2, max_row=365)
    c2.set_categories(dates)    
    
#     ws.add_chart(c2, "B%d" % (nrows+2))
    ws1 = wb['净值走势图']
    ws1.add_chart(c2, "B3")
    
    wb.save(filename) 
#     os.system(filename) #打开excel文件手工确认数据
    
    return 0


def sendmail(filename = 'c:/1.xlsx', receiver = ['iyangpingyan@icloud.com'], ccreceiver = [], subject="擎天柱基金资产净值明细表", pwfile='./emailpanna.json'):   
    mailinfo = file2dict(pwfile)
    sender = mailinfo['sender']
    smtpserver = mailinfo['smtpserver']
    username = mailinfo['username']
    password = mailinfo['password']
        
    # Create message container - the correct MIME type is multipart/alternative.    
    msg = MIMEMultipart('alternative')   
    msg['Subject']=subject
    msg['From']=sender 
    msg['To']=';'.join(receiver) 
    msg['Cc']=';'.join(ccreceiver) 
    #定义发送时间（不定义的可能有的邮件客户端会不显示发送时间） 
    msg['date']=time.strftime('%a, %d %b %Y %H:%M:%S %z')  
        
    # Create the body of the message (a plain-text and an HTML version).    
    text = "你好，\n    附件是擎天柱基金资产净值明细表，请查收。"    
    html = """\  
    <html>  
      <head></head>  
      <body>  
        <p>Hi!<br>  
           How are you?<br>  
           Here is the <a href="http://www.python.org">link</a> you wanted.  
        </p>  
      </body>  
    </html>  
    """    
        
    # Record the MIME types of both parts - text/plain and text/html.    
    part1 = MIMEText(text, 'plain')    
    part2 = MIMEText(html, 'html')    
    msg.attach(part1)    
    # msg.attach(part2)    

    basename = os.path.basename(filename)    
    ctype, encoding = mimetypes.guess_type(filename)
    if ctype is None or encoding is not None:
        # No guess could be made, or the file is encoded (compressed), so
        # use a generic bag-of-bits type.
        ctype = 'application/octet-stream'
        print("none.......")
    maintype, subtype = ctype.split('/', 1)
    fp = open(filename, 'rb')
    part3 = MIMEBase(maintype, subtype)
    part3.set_payload(fp.read())   
    encoders.encode_base64(part3)   
    part3.add_header('Content-Disposition', 'attachment',filename=('gbk', '', basename))
    msg.attach(part3)                   
            
    try:       
        smtp = smtplib.SMTP()    
        smtp.connect(smtpserver)    
        smtp.login(username, password)    
        smtp.sendmail(sender, receiver + ccreceiver, msg.as_string())    
        smtp.quit()  
        
    except Exception as e:
        print(str(e))
                

def parse_args(pargs=None):
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description='update details of fund qiantang1002')

    parser.add_argument('--assets', required=False, action='store',
                        type=float, default=0,          
                        help=('total assets of fund qiantang1002'))
    
    if pargs is not None:
        return parser.parse_args(pargs)

    return parser.parse_args()


    
if __name__ == '__main__':
    print("MISSION START")
    fundunit = 248450.07
    filename = './擎天柱基金资产净值明细表.xlsx'   
    pwfile = './emailpanna.json'     

    args = None
    args = parse_args(args)
    if args.assets:
        assets = args.assets
    else:
        str = input("基金净资产是：")
#        str = '192200'
        assets = float(str)
    print("assets={:.2f}, fundunit={:.2f}".format(assets,fundunit))

    hs300price = getprice_hs300()    
    
    retexcel = modify1002excel(filename = filename, assets=assets, hs300price=float(hs300price), fundunit=fundunit)
    
    subject="擎天柱基金今日净值{:.4f}".format(assets/fundunit)
    print(subject)
    
    if assets and (retexcel is 0):
        sendmail(filename = filename,receiver = ['joeyyh@126.com'], ccreceiver = ['iyangpingyan@icloud.com'], subject=subject)
        print("sent email success")
        pass
    else:
        print("no email sent")
        
      
    print("MISSION COMPLETE")


